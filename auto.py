import openpyxl
import github
from github import Github
import datetime

# GitHub username and personal access token
USERNAME = 'ZeroMemoryEx'
ACCESS_TOKEN = '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'

# Excel file path
EXCEL_FILE_PATH = 'C:/Users/Ajaybalaji/Downloads/ajay.xlsx'

# Worksheet name
WORKSHEET_NAME = 'automation'

# GitHub API client
g = Github(ACCESS_TOKEN)

# Load the Excel file
wb = openpyxl.load_workbook(EXCEL_FILE_PATH)

# Select the worksheet
ws = wb[WORKSHEET_NAME]

# Find the last row with data in the worksheet
last_row = ws.max_row

# Get the current date and time
now = datetime.datetime.now()

# Retrieve the list of repositories in the user's profile
repos = g.get_user(USERNAME).get_repos()

# Loop through the repositories and add them to the Excel sheet
for repo in repos:
    # Check if the repository is already in the Excel sheet
    found = False
    for row in range(2, last_row+1):
        if ws.cell(row=row, column=1).value == repo.name:
            found = True
            break
    # If the repository is not in the Excel sheet, add it
    if not found:
        last_row += 1
        ws.cell(row=last_row, column=1).value = repo.name
        ws.cell(row=last_row, column=2).value = repo.description
        ws.cell(row=last_row, column=3).value = now.strftime('%Y-%m-%d %H:%M:%S')

# Save the updated Excel file
wb.save(EXCEL_FILE_PATH)